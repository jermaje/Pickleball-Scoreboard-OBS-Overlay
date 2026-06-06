#!/usr/bin/env python3
"""
Simple HTTP server that serves the static scoreboard files and provides
an endpoint to read/update the scoreboard state at /state.

Usage:
  python server.py 8000 5

This avoids relying on browser localStorage so OBS and the control page
can share the same central state.
"""
import http.server
import socketserver
import json
import os
import sys
import threading
import time
import datetime
from urllib.parse import urlparse

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    print('Error: openpyxl is required to save state in Excel format.')
    print('Install it with: python -m pip install openpyxl')
    sys.exit(1)

STATE_FILE = 'state.xlsx'
STATE_SHEET = 'state'
STATE_HEADERS = ['team1Score', 'team2Score', 'team1Name', 'team2Name', 'serving', 'servingNumber', 'matchConfig', 'team1Sets', 'team2Sets', 'category', 'lastUpdated']
STATE_LOCK = threading.Lock()
AUTOSAVE_INTERVAL_SECONDS = 5
stop_event = threading.Event()
current_state = None

DEFAULT_STATE = {
    'team1Score': 0,
    'team2Score': 0,
    'team1Name': 'FIRST PLAYER',
    'team2Name': 'SECOND PLAYER',
    'serving': 1,
    'servingNumber': 1,
    'matchConfig': {
        'bestOf': 1,
        'pointsToWin': 11,
        'doubles': False
    },
    'team1Sets': 0,
    'team2Sets': 0,
    'category': 'MIXED DOUBLES',
    'lastUpdated': 0
}


def _create_state_backup(path):
    backup_path = f"{path}.bak"
    idx = 1
    while os.path.exists(backup_path):
        backup_path = f"{path}.bak.{idx}"
        idx += 1
    os.replace(path, backup_path)
    return backup_path


def load_state():
    try:
        wb = load_workbook(STATE_FILE, data_only=True)
        if STATE_SHEET not in wb.sheetnames:
            raise ValueError('Excel file missing required sheet')
        ws = wb[STATE_SHEET]

        headers = [cell.value for cell in ws[1][: len(STATE_HEADERS)]]
        valid_headers = [h for h in headers if h in STATE_HEADERS]
        if not valid_headers or set(valid_headers) < set(STATE_HEADERS[:-1]):
            raise ValueError('Excel sheet headers are invalid or missing')

        values = [ws.cell(row=2, column=i + 1).value for i in range(len(valid_headers))]
        raw_state = dict(zip(valid_headers, values))

        state = DEFAULT_STATE.copy()
        try:
            if raw_state.get('team1Score') is not None:
                state['team1Score'] = int(raw_state['team1Score'])
        except (TypeError, ValueError):
            pass
        try:
            if raw_state.get('team2Score') is not None:
                state['team2Score'] = int(raw_state['team2Score'])
        except (TypeError, ValueError):
            pass
        if raw_state.get('team1Name') is not None:
            state['team1Name'] = str(raw_state['team1Name'])
        if raw_state.get('team2Name') is not None:
            state['team2Name'] = str(raw_state['team2Name'])
        if raw_state.get('category') is not None:
            state['category'] = str(raw_state['category'])
        try:
            if raw_state.get('serving') is not None:
                serving_value = int(raw_state['serving'])
                if serving_value in (1, 2):
                    state['serving'] = serving_value
        except (TypeError, ValueError):
            pass
        try:
            if raw_state.get('servingNumber') is not None:
                serving_num_val = int(raw_state['servingNumber'])
                if serving_num_val in (1, 2):
                    state['servingNumber'] = serving_num_val
        except (TypeError, ValueError):
            pass
        try:
            if raw_state.get('team1Sets') is not None:
                state['team1Sets'] = int(raw_state['team1Sets'])
        except (TypeError, ValueError):
            pass
        try:
            if raw_state.get('team2Sets') is not None:
                state['team2Sets'] = int(raw_state['team2Sets'])
        except (TypeError, ValueError):
            pass
        try:
            if raw_state.get('lastUpdated') is not None:
                state['lastUpdated'] = int(raw_state['lastUpdated'])
        except (TypeError, ValueError):
            pass
        try:
            if raw_state.get('matchConfig') is not None:
                match_cfg = json.loads(raw_state['matchConfig']) if isinstance(raw_state['matchConfig'], str) else raw_state['matchConfig']
                if isinstance(match_cfg, dict):
                    state['matchConfig'] = match_cfg
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
        return state
    except FileNotFoundError:
        return DEFAULT_STATE.copy()
    except (InvalidFileException, ValueError) as e:
        try:
            backup_path = _create_state_backup(STATE_FILE)
            print(f'Warning: invalid state file {STATE_FILE}; moved broken file to {backup_path}: {e}')
        except Exception as backup_error:
            print(f'Warning: invalid state file {STATE_FILE}; could not create backup: {backup_error}')
        return DEFAULT_STATE.copy()
    except Exception as e:
        print(f'Warning: unable to load state file {STATE_FILE}: {e}')
        return DEFAULT_STATE.copy()


def save_state(state):
    wb = Workbook()
    ws = wb.active
    ws.title = STATE_SHEET
    ws.append(STATE_HEADERS)
    row = []
    for key in STATE_HEADERS:
        if key == 'matchConfig':
            val = state.get(key, DEFAULT_STATE[key])
            row.append(json.dumps(val) if isinstance(val, dict) else val)
        else:
            row.append(state.get(key, DEFAULT_STATE[key]))
    ws.append(row)
    wb.save(STATE_FILE)


def append_match_history(state):
    try:
        match_file = 'MATCH.xlsx'
        # Load existing workbook or create a new one
        if os.path.exists(match_file):
            wb = load_workbook(match_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Match History"
            ws.append(['Timestamp', 'Category', 'Team 1 Name', 'Team 1 Score', 'Team 2 Name', 'Team 2 Score', 'Winner'])
        
        t1_score = state.get('team1Score', 0)
        t2_score = state.get('team2Score', 0)
        t1_name = state.get('team1Name', 'FIRST PLAYER')
        t2_name = state.get('team2Name', 'SECOND PLAYER')
        category = state.get('category', 'MIXED DOUBLES')
        
        if t1_score > t2_score:
            winner = t1_name
        elif t2_score > t1_score:
            winner = t2_name
        else:
            winner = 'TIE'
            
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append([timestamp, category, t1_name, t1_score, t2_name, t2_score, winner])
        wb.save(match_file)
        return True, None
    except Exception as e:
        return False, str(e)


def load_match_history():
    match_file = 'MATCH.xlsx'
    if not os.path.exists(match_file):
        return []
    try:
        # Load workbook (read-only for speed)
        wb = load_workbook(match_file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            return []
        
        headers = rows[0]
        history = []
        for row in rows[1:]:
            if not any(row):  # skip empty rows
                continue
            # Map row values to headers
            row_dict = {}
            for idx, key in enumerate(headers):
                if idx < len(row):
                    row_dict[key] = row[idx]
                else:
                    row_dict[key] = None
            history.append(row_dict)
        return history
    except Exception as e:
        print(f"Error loading match history: {e}")
        return []


def autosave_loop():
    while not stop_event.wait(AUTOSAVE_INTERVAL_SECONDS):
        with STATE_LOCK:
            save_state(current_state)


def notify_shutdown():
    stop_event.set()
    with STATE_LOCK:
        save_state(current_state)


def validate_update(data):
    valid = {}
    for key, value in data.items():
        if key == 'team1Score' or key == 'team2Score':
            if isinstance(value, int) and value >= 0:
                valid[key] = value
            else:
                return None, f'{key} must be a non-negative integer'
        elif key == 'serving':
            if value in (1, 2):
                valid[key] = value
            else:
                return None, 'serving must be 1 or 2'
        elif key == 'servingNumber':
            if value in (1, 2):
                valid[key] = value
            else:
                return None, 'servingNumber must be 1 or 2'
        elif key in ('team1Sets', 'team2Sets'):
            if isinstance(value, int) and value >= 0:
                valid[key] = value
            else:
                return None, f'{key} must be a non-negative integer'
        elif key in ('team1Name', 'team2Name', 'category'):
            if isinstance(value, str):
                valid[key] = value
            else:
                return None, f'{key} must be a string'
        elif key == 'matchConfig':
            if isinstance(value, dict):
                # Validate matchConfig structure
                match_cfg = {}
                if 'bestOf' in value:
                    if value['bestOf'] in (1, 3, 5):
                        match_cfg['bestOf'] = value['bestOf']
                    else:
                        return None, 'matchConfig.bestOf must be 1, 3, or 5'
                else:
                    match_cfg['bestOf'] = DEFAULT_STATE['matchConfig']['bestOf']
                if 'pointsToWin' in value:
                    if value['pointsToWin'] in (11, 15, 21):
                        match_cfg['pointsToWin'] = value['pointsToWin']
                    else:
                        return None, 'matchConfig.pointsToWin must be 11, 15, or 21'
                else:
                    match_cfg['pointsToWin'] = DEFAULT_STATE['matchConfig']['pointsToWin']
                if 'doubles' in value:
                    if isinstance(value['doubles'], bool):
                        match_cfg['doubles'] = value['doubles']
                    else:
                        return None, 'matchConfig.doubles must be a boolean'
                else:
                    match_cfg['doubles'] = DEFAULT_STATE['matchConfig']['doubles']
                valid[key] = match_cfg
            else:
                return None, 'matchConfig must be an object'
        else:
            return None, f'Unsupported field: {key}'
    return valid, None


class ScoreHandler(http.server.SimpleHTTPRequestHandler):
    def _send_json(self, obj, code=200):
        data = json.dumps(obj).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(data)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path in ('/state', '/state.json'):
            with STATE_LOCK:
                self._send_json(current_state)
            return
        elif parsed.path in ('/history-data', '/history_data'):
            history = load_match_history()
            self._send_json(history)
            return
        return super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == '/state':
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length) if length else b''
            try:
                data = json.loads(body.decode('utf-8')) if body else {}
            except Exception:
                self._send_json({'error': 'invalid json'}, 400)
                return

            if not isinstance(data, dict):
                self._send_json({'error': 'payload must be a JSON object'}, 400)
                return

            valid_update, error = validate_update(data)
            if error:
                self._send_json({'error': error}, 400)
                return

            with STATE_LOCK:
                current_state.update(valid_update)
                current_state['lastUpdated'] = int(time.time() * 1000)
                updated = current_state.copy()

            self._send_json(updated)
            return
        elif parsed.path == '/save_match':
            with STATE_LOCK:
                success, error_msg = append_match_history(current_state)
            if success:
                self._send_json({'status': 'success'})
            else:
                self._send_json({'error': error_msg or 'unknown error'}, 500)
            return

        self.send_response(404)
        self.end_headers()


def ensure_state_file():
    if not os.path.exists(STATE_FILE):
        try:
            save_state(DEFAULT_STATE.copy())
        except Exception as e:
            print(f'Warning: unable to create state file {STATE_FILE}: {e}')


def run(port=8000):
    global current_state
    current_state = load_state()
    ensure_state_file()
    save_state(current_state)
    autosave_thread = threading.Thread(target=autosave_loop, daemon=True)
    autosave_thread.start()

    socketserver.ThreadingTCPServer.allow_reuse_address = True
    handler = ScoreHandler
    with socketserver.ThreadingTCPServer(('', port), handler) as httpd:
        print(f"Serving on http://0.0.0.0:{port}")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print('\nShutting down')
            notify_shutdown()
            httpd.server_close()


if __name__ == '__main__':
    port = 8000
    autosave_interval = AUTOSAVE_INTERVAL_SECONDS
    if len(sys.argv) > 1:
        try:
            port = int(sys.argv[1])
        except Exception:
            pass
    if len(sys.argv) > 2:
        try:
            autosave_interval = int(sys.argv[2])
            if autosave_interval < 1:
                raise ValueError('interval must be >= 1')
        except Exception:
            print('Warning: invalid autosave interval, using default', AUTOSAVE_INTERVAL_SECONDS)
            autosave_interval = AUTOSAVE_INTERVAL_SECONDS
    AUTOSAVE_INTERVAL_SECONDS = autosave_interval
    print(f'Autosave interval: {AUTOSAVE_INTERVAL_SECONDS} seconds')
    run(port)
